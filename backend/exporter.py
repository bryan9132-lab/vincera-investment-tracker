"""Generate 庫存總表 Excel — RC/華強 merged (no 私RC/私強ニ split), no highlight colors"""
import tempfile
from datetime import date, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from .models import db, Position, Security, CashAccount, Transaction
from .logic import calculate_realized_pnl

# Map detailed entity → display group (RC+私銀RC merge into RC, 華強+私銀華強 merge into 華強)
ENTITY_GROUP = {'RC': 'RC', '私銀RC': 'RC', '華強': '華強', '私銀華強': '華強'}
DISPLAY_ENTITIES = ['RC', '華強']


def _thin_border():
    t = Side(style='thin', color='000000')
    return Border(left=t, right=t, top=t, bottom=t)


def _apply_range_border(ws, min_row, min_col, max_row, max_col):
    """Apply thin outer border to every edge cell of a merged or plain range."""
    t = Side(style='thin', color='000000')
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(r, c)
            left   = t if c == min_col else None
            right  = t if c == max_col else None
            top    = t if r == min_row else None
            bottom = t if r == max_row else None
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)


def _cell(ws, row, col, value=None, font_size=20, bold=False,
          align='center', valign='center', wrap=False,
          num_fmt=None, color='000000', border=False):
    c = ws.cell(row, col, value)
    c.font      = Font(name='微軟正黑體', size=font_size, bold=bold, color=color)
    c.alignment = Alignment(horizontal=align, vertical=valign, wrap_text=wrap)
    if num_fmt: c.number_format = num_fmt
    if border:  c.border = _thin_border()
    return c


def generate_excel() -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = '庫存總表(台灣)'

    # ── Page setup ────────────────────────────────────────────────────────────
    ws.page_setup.paperSize   = 9  # A4
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=2/2.54, right=2/2.54,
                                  top=2.3/2.54, bottom=2.3/2.54,
                                  header=0.8/2.54, footer=0.8/2.54)

    # ── Column widths ────────────────────────────────────────────────────────
    col_w = {
        'A': 13, 'B': 21, 'C': 24, 'D': 20,
        'E': 10, 'F': 14, 'G': 24,
        'H': 10, 'I': 14, 'J': 24,
        'K': 17, 'L': 16,
    }
    for col, w in col_w.items():
        ws.column_dimensions[col].width = w

    today_str = date.today().strftime('%Y/%m/%d')

    # ── Row 1: Title ──────────────────────────────────────────────────────────
    ws.merge_cells('A1:L1')
    _cell(ws, 1, 1, f'庫存總表（台灣）　　單位：NTD　　日期：{today_str}',
          font_size=28, bold=True, align='center')
    ws.row_dimensions[1].height = 88.5

    # ── Row 2: 今日投資總結 (no highlight, just a text block) ──────────────────
    today = date.today()
    todays_txns = (Transaction.query
                  .filter(Transaction.trade_date == today)
                  .filter(Transaction.security_code.isnot(None))
                  .order_by(Transaction.entity, Transaction.broker)
                  .all())
    summary_lines = ['今日投資總結：']
    if todays_txns:
        for t in todays_txns:
            zhang = abs(t.shares) / 1000
            zhang_str = f'{zhang:.3f}'.rstrip('0').rstrip('.')
            action = '買進' if t.shares > 0 else '賣出'
            summary_lines.append(
                f'{t.entity} {t.broker} {action} {t.security_name or t.security_code} '
                f'{zhang_str}張 @ {t.price:.2f}'
            )
    else:
        summary_lines.append('（今日無投資交易）')

    ws.merge_cells('A2:L2')
    _cell(ws, 2, 1, '\n'.join(summary_lines), font_size=16, bold=False,
          align='left', valign='top', wrap=True)
    ws.row_dimensions[2].height = max(74.25, 22 * len(summary_lines))

    # ── Row 3-4: Group headers ───────────────────────────────────────────────
    for col in ['A', 'B', 'C', 'D']:
        ws.merge_cells(f'{col}3:{col}4')
    headers_r3 = [('A3','A4','股票\n代號'), ('B3','B4','股票名稱'), ('C3','C4','券商'),
                  ('D3','D4','市價'), ('E3','G3','RC'), ('H3','J3','華強'), ('K3','L3','未實現損益')]
    for start, end, label in headers_r3:
        if start != end:
            ws.merge_cells(f'{start}:{end}')
        sc = ws[start]
        sc.value     = label
        sc.font      = Font(name='微軟正黑體', size=16, bold=True)
        sc.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        sc.border    = _thin_border()
    ws.row_dimensions[3].height = 36  # locked

    sub_headers = {5:'張數', 6:'成本', 7:'金額', 8:'張數', 9:'成本', 10:'金額', 11:'RC', 12:'華強'}
    for col_i, label in sub_headers.items():
        c = ws.cell(4, col_i, label)
        c.font      = Font(name='微軟正黑體', size=16, bold=True)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border    = _thin_border()
    ws.row_dimensions[4].height = 51  # locked

    # ── Data rows — merge RC+私銀RC, 華強+私銀華強 ──────────────────────────────
    positions = Position.query.filter(Position.shares > 0).all()
    by_code = {}
    for pos in positions:
        code = pos.security_code
        disp_ent = ENTITY_GROUP.get(pos.entity, pos.entity)
        if code not in by_code:
            sec = Security.query.get(code)
            by_code[code] = {'name': sec.name if sec else code, 'price': pos.last_price,
                             'groups': {}}
        if pos.last_price:
            by_code[code]['price'] = pos.last_price
        g = by_code[code]['groups'].setdefault(disp_ent, {'shares': 0.0, 'total_cost': 0.0})
        g['shares']     += pos.shares
        g['total_cost'] += pos.total_cost

    # 券商 per stock — most recent transaction's broker
    broker_by_code = {}
    all_txns = (Transaction.query
               .filter(Transaction.security_code.isnot(None))
               .order_by(Transaction.trade_date.desc())
               .all())
    for t in all_txns:
        if t.security_code not in broker_by_code:
            broker_by_code[t.security_code] = t.broker

    realized = calculate_realized_pnl()
    row = 5
    totals  = {e: 0 for e in DISPLAY_ENTITIES}
    tot_pnl = {e: 0 for e in DISPLAY_ENTITIES}

    ROW_HEIGHTS = [67.5, 76.5, 69.75, 81.75, 84.0, 67.5]  # per-stock row heights

    for idx, (code, data) in enumerate(sorted(by_code.items())):
        price  = data['price']
        broker = broker_by_code.get(code, '')

        _cell(ws, row, 1, code,         font_size=20, bold=True, align='center', border=True)
        _cell(ws, row, 2, data['name'], font_size=20, bold=True, align='left',   border=True)
        _cell(ws, row, 3, broker,       font_size=20, bold=True, align='center', border=True)
        _cell(ws, row, 4, price,        font_size=20, align='right', border=True, num_fmt='#,##0.00')

        col = 5
        pnl_vals = {}
        for ent in DISPLAY_ENTITIES:
            g = data['groups'].get(ent)
            shares = (g['shares'] / 1000) if g else None
            if shares is not None:
                shares = float(f'{shares:.3f}'.rstrip('0').rstrip('.'))
            cost  = (g['total_cost'] / g['shares']) if (g and g['shares']) else None
            total = round(g['total_cost']) if g else None
            pnl   = round((price * g['shares']) - g['total_cost']) if (g and price) else None
            pnl_vals[ent] = pnl

            _cell(ws, row, col,   shares, font_size=20, align='right', border=True, num_fmt='#,##0.##')
            _cell(ws, row, col+1, cost,   font_size=20, align='right', border=True, num_fmt='#,##0.00')
            _cell(ws, row, col+2, total,  font_size=20, align='right', border=True, num_fmt='#,##0')
            if total: totals[ent] += total
            col += 3

        for i, ent in enumerate(DISPLAY_ENTITIES):
            pnl = pnl_vals.get(ent)
            c = ws.cell(row, 11 + i, pnl)
            clr = 'C00000' if (pnl or 0) < 0 else '000000'
            c.font          = Font(name='微軟正黑體', size=20, color=clr)
            c.alignment     = Alignment(horizontal='right', vertical='center')
            c.number_format = '#,##0'
            c.border        = _thin_border()
            if pnl: tot_pnl[ent] += pnl

        ws.row_dimensions[row].height = ROW_HEIGHTS[idx % len(ROW_HEIGHTS)]
        row += 1

    # ── Totals row ────────────────────────────────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    _cell(ws, row, 1, '小計', font_size=20, bold=True, align='left', border=True)
    ws.cell(row, 4).border = _thin_border()
    col = 5
    for ent in DISPLAY_ENTITIES:
        _cell(ws, row, col,   None,        font_size=20, border=True)
        _cell(ws, row, col+1, None,        font_size=20, border=True)
        _cell(ws, row, col+2, totals[ent], font_size=20, bold=True, align='right',
              border=True, num_fmt='#,##0')
        col += 3
    for i, ent in enumerate(DISPLAY_ENTITIES):
        val = tot_pnl[ent] or None
        c = ws.cell(row, 11+i, val)
        clr = 'C00000' if (val or 0) < 0 else '000000'
        c.font          = Font(name='微軟正黑體', size=20, bold=True, color=clr)
        c.alignment     = Alignment(horizontal='right', vertical='center')
        c.number_format = '#,##0'
        c.border        = _thin_border()
    ws.row_dimensions[row].height = 51  # 小計 row
    row += 2
    ws.row_dimensions[row-1].height = 15.75  # blank spacer row 7

    # ── 損益 + 資金餘額 side-by-side ────────────────────────────────────────
    rRC  = realized.get('RC', 0)   + realized.get('私銀RC', 0)
    rHQ  = realized.get('華強', 0) + realized.get('私銀華強', 0)
    uRC  = sum((p.unrealized_pnl() or 0) for p in Position.query
              .filter(Position.entity.in_(['RC','私銀RC'])).filter(Position.shares>0).all())
    uHQ  = sum((p.unrealized_pnl() or 0) for p in Position.query
              .filter(Position.entity.in_(['華強','私銀華強'])).filter(Position.shares>0).all())

    RC_PRIVATE  = {'rc_private'}
    HQ_PRIVATE  = {'hq_private'}
    RC_NON_PRIV = {'rc_dunnan','rc_tuni','rc_yuanta','rc_fund','rc_other'}
    HQ_NON_PRIV = {'hq_tuni','hq_yuanta','hq_dunnan','hq_fund','hq_huanan','hq_fubon','hq_yuanta_bank'}
    rc_balance = 0
    hq_balance = 0
    for acct in CashAccount.query.all():
        bal = acct.balance or 0
        if acct.id in RC_NON_PRIV or acct.id in RC_PRIVATE:
            rc_balance += bal
        elif acct.id in HQ_NON_PRIV or acct.id in HQ_PRIVATE:
            hq_balance += bal

    pnl_start = row

    # ── LEFT: 損益 (cols A-D, rows pnl_start to pnl_start+4) ────────────────
    ws.merge_cells(start_row=pnl_start, start_column=1, end_row=pnl_start+1, end_column=2)
    _cell(ws, pnl_start, 1, '損益', font_size=16, bold=True, align='center')
    _apply_range_border(ws, pnl_start, 1, pnl_start+1, 2)
    ws.merge_cells(start_row=pnl_start, start_column=3, end_row=pnl_start, end_column=4)
    _cell(ws, pnl_start, 3, '合計', font_size=16, bold=True, align='center')
    _apply_range_border(ws, pnl_start, 3, pnl_start, 4)
    ws.row_dimensions[pnl_start].height = 59.25

    for col_i, lbl in [(3,'RC'), (4,'華強')]:
        c2 = ws.cell(pnl_start+1, col_i, lbl)
        c2.font      = Font(name='微軟正黑體', size=16, bold=True)
        c2.alignment = Alignment(horizontal='center', vertical='center')
        c2.border    = _thin_border()
    ws.row_dimensions[pnl_start+1].height = 67.5

    pnl_data = [
        ('已實現損益', rRC, rHQ),
        ('未實現損益', uRC, uHQ),
        ('合計', rRC+uRC, rHQ+uHQ),
    ]
    pnl_row_heights = [77.25, 69, 55.5]
    for i, (label, rc, hq) in enumerate(pnl_data):
        r = pnl_start + 2 + i
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        _cell(ws, r, 1, label, font_size=16, bold=True, align='left')
        _apply_range_border(ws, r, 1, r, 2)
        for col_i, val in enumerate([rc, hq], 3):
            clr = 'C00000' if val < 0 else '000000'
            c = ws.cell(r, col_i, round(val))
            c.font          = Font(name='微軟正黑體', size=16, bold=(label=='合計'), color=clr)
            c.alignment     = Alignment(horizontal='right', vertical='center')
            c.number_format = '#,##0'
            c.border        = _thin_border()
        ws.row_dimensions[r].height = pnl_row_heights[i]

    # ── RIGHT: 資金餘額 (cols F-I, rows pnl_start to pnl_start+2) ────────────
    ws.merge_cells(start_row=pnl_start, start_column=6, end_row=pnl_start, end_column=7)
    _cell(ws, pnl_start, 6, '賬戶', font_size=16, bold=True, align='center')
    _apply_range_border(ws, pnl_start, 6, pnl_start, 7)
    ws.merge_cells(start_row=pnl_start, start_column=8, end_row=pnl_start, end_column=9)
    _cell(ws, pnl_start, 8, '資金餘額', font_size=16, bold=True, align='center')
    _apply_range_border(ws, pnl_start, 8, pnl_start, 9)

    for i, (label, bal) in enumerate([('RC', rc_balance), ('華強', hq_balance)]):
        r = pnl_start + 1 + i
        ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=7)
        _cell(ws, r, 6, label, font_size=16, bold=True, align='left')
        _apply_range_border(ws, r, 6, r, 7)
        ws.merge_cells(start_row=r, start_column=8, end_row=r, end_column=9)
        c = ws.cell(r, 8, round(bal))
        c.font          = Font(name='微軟正黑體', size=20)
        c.alignment     = Alignment(horizontal='right', vertical='center')
        c.number_format = '#,##0'
        _apply_range_border(ws, r, 8, r, 9)

    row = pnl_start + 5

    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False,
        prefix=f'庫存總表_{date.today().strftime("%Y%m%d")}_')
    wb.save(tmp.name)
    return tmp.name
