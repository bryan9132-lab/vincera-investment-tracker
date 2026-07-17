"""
Flask API — all routes Sophie's frontend will call
"""

import os
import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from datetime import datetime, date
import tempfile

from backend.models import (db, Transaction, Position, Security, UploadLog,
                            ACCOUNT_MAP, CASH_ACCOUNTS, CASH_ACCOUNT_BROKER_MAP,
                            FUND_CASH_LINK, CASH_ACCOUNTS_BY_ID,
                            CashAccount, CashEntry, FundEntry, AuditLog,
                            CashDividend, StockDividend, DIVIDEND_BROKERS, DIVIDEND_ENTITIES)
from backend.logic  import recalculate_positions, update_all_prices, fetch_twse_name, calculate_realized_pnl
from parsers.pdf_parsers import parse_pdf
from backend.exporter import generate_excel


def create_app():
    app = Flask(__name__, static_folder='../frontend', static_url_path='')
    CORS(app)

    # Config
    # Railway provides postgres:// but SQLAlchemy needs postgresql://
    # Try multiple possible env variable names Railway might use
    db_url = (
        os.environ.get('DATABASE_URL') or
        os.environ.get('DATABASE_PRIVATE_URL') or
        os.environ.get('POSTGRES_URL') or
        os.environ.get('POSTGRESQL_URL') or
        'sqlite:///sophie.db'
    )
    if db_url.startswith('postgres://'):
        db_url = db_url.replace('postgres://', 'postgresql://', 1)
    # Log which DB we're using (visible in Railway logs)
    import logging
    logging.warning(f'VIT: Using database: {db_url[:30]}...')
    app.config['SQLALCHEMY_DATABASE_URI'] = db_url
    app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
    app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max upload

    db.init_app(app)

    with app.app_context():
        db.create_all()
        _migrate_add_price_type()
        _seed_securities()
        _seed_cash_accounts()

    # ── Serve frontend ───────────────────────────────────────────────────────
    @app.route('/')
    def index():
        return app.send_static_file('index.html')

    # ── PDF Upload & Parse ───────────────────────────────────────────────────
    @app.route('/api/upload', methods=['POST'])
    def upload_pdf():
        """
        Step 1: Upload PDF, parse it, return preview for Sophie to confirm.
        Does NOT save to DB yet.
        """
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400

        file       = request.files['file']
        pdf_bytes  = file.read()
        filename   = file.filename

        parsed = parse_pdf(pdf_bytes, filename)

        if parsed.broker == 'unknown' or not parsed.account_no:
            return jsonify({
                'error': 'Could not identify broker or account',
                'details': parsed.errors
            }), 422

        # Check account mapping
        mapping = ACCOUNT_MAP.get(parsed.account_no)
        if not mapping:
            return jsonify({
                'error': f'Unknown account number: {parsed.account_no}',
                'details': ['This account is not in our mapping. Please contact admin.']
            }), 422

        # Check for duplicate upload
        force    = request.form.get('force', 'false').lower() == 'true'
        existing = UploadLog.query.filter_by(
            account_no = parsed.account_no,
            trade_date = parsed.trade_date,
        ).first()
        if existing and existing.status == 'confirmed' and not force:
            return jsonify({
                'warning': 'duplicate',
                'message': f'{parsed.account_no} 在 {parsed.trade_date} 的交易已上傳過。確定要再次上傳嗎？',
                'account_no':  parsed.account_no,
                'entity':      mapping['entity'],
                'broker':      parsed.broker,
                'trade_date':  parsed.trade_date.isoformat(),
                'settle_date': parsed.settle_date.isoformat() if parsed.settle_date else None,
                'trades':      [],
                'unknown_codes': [],
                'warnings':    [],
            }), 200

        # Enrich with known security names
        trades_preview = []
        unknown_codes  = []
        for t in parsed.trades:
            sec  = Security.query.get(t.security_code)
            name = sec.name if sec else fetch_twse_name(t.security_code)
            if not name:
                unknown_codes.append(t.security_code)
            trades_preview.append({
                'security_code': t.security_code,
                'security_name': name or f'[Unknown: {t.security_code}]',
                'name_confirmed': bool(name),
                'shares':        t.shares,
                'price':         t.price,
                'gross_amount':  t.gross_amount,
                'fee':           t.fee,
                'tax':           t.tax,
                'net_amount':    t.net_amount,
            })

        return jsonify({
            'account_no':    parsed.account_no,
            'entity':        mapping['entity'],
            'broker':        parsed.broker,
            'trade_date':    parsed.trade_date.isoformat(),
            'settle_date':   parsed.settle_date.isoformat() if parsed.settle_date else None,
            'trades':        trades_preview,
            'unknown_codes': unknown_codes,
            'warnings':      parsed.errors,
        })

    # ── Confirm parsed PDF ───────────────────────────────────────────────────
    @app.route('/api/confirm', methods=['POST'])
    def confirm_upload():
        """
        Step 2: Sophie reviews and confirms — save to DB.
        Accepts the same structure returned by /api/upload,
        with any name corrections applied.
        """
        data = request.json
        if not data:
            return jsonify({'error': 'No data'}), 400

        account_no  = data['account_no']
        trade_date  = date.fromisoformat(data['trade_date'])
        settle_date = date.fromisoformat(data['settle_date']) if data.get('settle_date') else None
        mapping     = ACCOUNT_MAP[account_no]
        trades      = data['trades']

        # Save/update securities
        for t in trades:
            code = t['security_code']
            name = t['security_name']
            if code and not code.startswith('[Unknown'):
                sec = Security.query.get(code)
                if sec is None:
                    sec = Security(code=code, name=name)
                    db.session.add(sec)
                elif not sec.name or sec.name != name:
                    sec.name = name

        # Save transactions
        for t in trades:
            txn = Transaction(
                trade_date     = trade_date,
                settle_date    = settle_date,
                entity         = mapping['entity'],
                broker         = mapping['broker'],
                account_no     = account_no,
                security_code  = t['security_code'] if not t['security_code'].startswith('[') else None,
                security_name  = t['security_name'],
                shares         = t['shares'],
                price          = t['price'],
                gross_amount   = t['gross_amount'],
                fee            = t['fee'],
                tax            = t['tax'],
                net_amount     = t['net_amount'],
                source_file    = data.get('filename', ''),
            )
            db.session.add(txn)

        # Log upload
        log = UploadLog.query.filter_by(
            account_no=account_no, trade_date=trade_date
        ).first()
        if log is None:
            log = UploadLog(
                filename   = data.get('filename', ''),
                account_no = account_no,
                trade_date = trade_date,
                broker     = mapping['broker'],
            )
            db.session.add(log)
        log.status       = 'confirmed'
        log.confirmed_at = datetime.utcnow()

        db.session.commit()

        # Recalculate positions for this entity
        recalculate_positions(mapping['entity'])

        # ── Auto-update cash balance for investment account ───────────────
        cash_acct_id = CASH_ACCOUNT_BROKER_MAP.get(account_no)
        if cash_acct_id:
            cash_acct = CashAccount.query.get(cash_acct_id)
            if cash_acct and not cash_acct.is_static:
                saved_txns = Transaction.query.filter_by(
                    account_no = account_no,
                    trade_date = trade_date,
                ).all()
                for txn in saved_txns:
                    is_buy      = txn.shares > 0
                    cash_amount = -txn.net_amount if is_buy else txn.net_amount
                    shares_lots = int(abs(txn.shares) / 1000) if abs(txn.shares) >= 1000 else abs(txn.shares)
                    lots_unit   = '張' if abs(txn.shares) >= 1000 else '股'
                    per_share   = round(txn.price, 2)
                    desc = f"{txn.security_code} {txn.security_name or ''} {shares_lots}{lots_unit} @{per_share}/股"
                    cash_acct.balance += cash_amount
                    ce = CashEntry(
                        account_id     = cash_acct_id,
                        entry_date     = trade_date,
                        entry_type     = '買入股票' if is_buy else '賣出股票',
                        description    = desc.strip(),
                        amount         = cash_amount,
                        balance_after  = cash_acct.balance,
                        security_code  = txn.security_code,
                        security_name  = txn.security_name,
                        transaction_id = txn.id,
                        is_auto        = True,
                    )
                    db.session.add(ce)
                db.session.commit()

        return jsonify({'status': 'ok', 'message': f'Saved {len(trades)} transactions'})

    # ── Get current positions ────────────────────────────────────────────────
    @app.route('/api/positions', methods=['GET'])
    def get_positions():
        """Return all current positions grouped by entity, with pending 股票股利 attached"""
        positions = Position.query.filter(Position.shares > 0).all()
        result    = {}
        for pos in positions:
            ent = pos.entity
            if ent not in result:
                result[ent] = []
            result[ent].append(pos.to_dict())

        # Attach pending (not yet allocated) stock dividends to their entity+code
        pending = StockDividend.query.filter_by(allocated=False).all()
        for sd in pending:
            pos_entity = _dividend_position_entity(sd.entity, sd.broker)
            if pos_entity in result:
                for p in result[pos_entity]:
                    if p['security_code'] == sd.security_code:
                        p.setdefault('pending_stock_dividends', []).append({
                            'id': sd.id,
                            'bonus_shares': sd.bonus_shares,
                            'expected_allocate_date': sd.expected_allocate_date.isoformat() if sd.expected_allocate_date else None,
                        })
        return jsonify(result)


    # ── Realized P&L ────────────────────────────────────────────────────────
    @app.route('/api/realized_pnl', methods=['GET'])
    def get_realized_pnl():
        """Calculate realized P&L for all entities"""
        result = calculate_realized_pnl()
        return jsonify(result)

    @app.route('/api/realized_pnl_ledger', methods=['GET'])
    def get_realized_pnl_ledger():
        """Full audit-trail ledger for 已實現損益 tab — read-only."""
        from .logic import get_realized_pnl_ledger as _ledger
        rows = _ledger(
            entity   = request.args.get('entity')   or None,
            broker   = request.args.get('broker')   or None,
            category = request.args.get('category') or None,
        )
        return jsonify(rows)

    # ── Update market prices ─────────────────────────────────────────────────
    @app.route('/api/prices/update', methods=['POST'])
    def update_prices():
        """Fetch latest TWSE prices for all held securities"""
        results = update_all_prices()
        return jsonify({
            'updated': len([r for r in results if r['status'] == 'updated']),
            'failed':  len([r for r in results if r['status'] == 'failed']),
            'details': results,
        })

    # ── Export Excel ─────────────────────────────────────────────────────────
    @app.route('/api/export', methods=['GET'])
    def export_excel():
        """Generate and download the 庫存總表 Excel report"""
        filepath = generate_excel()
        return send_file(
            filepath,
            as_attachment=True,
            download_name=f'庫存總表_{date.today().strftime("%Y%m%d")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    # ── Manual transaction entry ────────────────────────────────────────────
    @app.route('/api/manual_entry', methods=['POST'])
    def manual_entry():
        """
        Preview manually entered transactions (same flow as PDF upload).
        Accepts JSON with account_no, trade_date, and trades list.
        """
        data       = request.json
        account_no = data.get('account_no', '').strip()
        trade_date_str = data.get('trade_date', '')
        trades_raw = data.get('trades', [])
        force      = data.get('force', False)

        mapping = ACCOUNT_MAP.get(account_no)
        if not mapping:
            return jsonify({'error': f'未知帳號：{account_no}'}), 422

        try:
            trade_date = date.fromisoformat(trade_date_str)
        except Exception:
            return jsonify({'error': '日期格式錯誤'}), 422

        # Duplicate check
        existing = UploadLog.query.filter_by(
            account_no=account_no,
            trade_date=trade_date,
        ).first()
        if existing and existing.status == 'confirmed' and not force:
            return jsonify({
                'warning': 'duplicate',
                'message': f'{account_no} 在 {trade_date} 的交易已上傳過。確定要再次上傳嗎？',
                'account_no': account_no, 'entity': mapping['entity'],
                'broker': mapping['broker'],
                'trade_date': trade_date.isoformat(), 'settle_date': None,
                'trades': [], 'unknown_codes': [], 'warnings': [],
            }), 200

        # Enrich trades with security names
        trades_preview = []
        unknown_codes  = []
        for t in trades_raw:
            code = str(t.get('security_code', '')).strip()
            sec  = Security.query.get(code)
            name = sec.name if sec else fetch_twse_name(code)
            if not name:
                unknown_codes.append(code)
                name = f'[Unknown: {code}]'
            elif sec and not sec.name:
                sec.name = name

            shares = float(t.get('shares', 0))
            price  = float(t.get('price', 0))
            gross  = abs(shares) * price
            FEE_RATES = {'統一': 0.00036477, '元大': 0.000625, '國泰': 0.00042465}
            fee_rate = FEE_RATES.get(mapping['broker'], 0.001425)
            fee    = round(gross * fee_rate)  # 小數點四捨五入
            tax    = round(gross * 0.003) if shares < 0 else 0
            net    = gross - fee - tax

            trades_preview.append({
                'security_code':  code,
                'security_name':  t.get('security_name', name),
                'name_confirmed': bool(sec and sec.name),
                'shares':         shares,
                'price':          price,
                'gross_amount':   round(gross),
                'fee':            int(t.get('fee', fee)),
                'tax':            int(t.get('tax', tax)),
                'net_amount':     round(net),
            })

        return jsonify({
            'account_no':    account_no,
            'entity':        mapping['entity'],
            'broker':        mapping['broker'],
            'trade_date':    trade_date.isoformat(),
            'settle_date':   None,
            'trades':        trades_preview,
            'unknown_codes': unknown_codes,
            'warnings':      [],
            'source':        'manual',
        })

    # ── One-time migration endpoint ──────────────────────────────────────────
    @app.route('/api/migrate', methods=['POST'])
    def run_migration():
        """
        One-time endpoint to migrate data from uploaded Excel.
        Protected by a secret key.
        """
        secret = request.json.get('secret', '')
        if secret != os.environ.get('MIGRATE_SECRET', 'vincera2026'):
            return jsonify({'error': 'Unauthorized'}), 401

        try:
            import openpyxl, sys
            from io import BytesIO
            import base64

            excel_b64 = request.json.get('excel_data')
            if not excel_b64:
                return jsonify({'error': 'No excel_data provided'}), 400

            excel_bytes = base64.b64decode(excel_b64)
            wb = openpyxl.load_workbook(BytesIO(excel_bytes), data_only=True)

            sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
            from migrate_excel import (migrate_rc_tuni, migrate_rc_yuanta,
                                       migrate_hq_tuni, migrate_hq_yuanta,
                                       migrate_private_rc, migrate_private_hq)

            # Clear existing
            Transaction.query.delete()
            Position.query.delete()
            db.session.commit()

            sheet_handlers = [
                ('(RC)統一進出明細TWD',   migrate_rc_tuni),
                ('(RC)元大進出明細TWD',   migrate_rc_yuanta),
                ('(華強)統一進出明細TWD', migrate_hq_tuni),
                ('(華強)元大進出明細TWD', migrate_hq_yuanta),
                ('(私RC)TWD',            migrate_private_rc),
                ('(私強)進出明細TWD',     migrate_private_hq),
            ]

            total = 0
            summary = {}
            for sheet_name, handler in sheet_handlers:
                ws     = wb[sheet_name]
                trades = handler(ws)
                summary[sheet_name] = len(trades)
                for t in trades:
                    code = t['security_code']
                    if code and not Security.query.get(code):
                        db.session.add(Security(code=code, name=code))
                    txn = Transaction(
                        trade_date    = t['trade_date'],
                        entity        = t['entity'],
                        broker        = t['broker'],
                        account_no    = t['account_no'],
                        security_code = t['security_code'],
                        shares        = t['shares'],
                        price         = t['price'],
                        gross_amount  = t['gross_amount'],
                        fee           = t['fee'],
                        tax           = t['tax'],
                        net_amount    = t['net_amount'],
                        source_file   = 'excel_migration',
                    )
                    db.session.add(txn)
                    total += 1

            db.session.commit()
            recalculate_positions()

            return jsonify({
                'status': 'ok',
                'total_transactions': total,
                'summary': summary,
            })

        except Exception as e:
            db.session.rollback()
            return jsonify({'error': str(e)}), 500

    # ── Transaction history ──────────────────────────────────────────────────
    @app.route('/api/transactions', methods=['GET'])
    def get_transactions():
        """Return recent transactions, optionally filtered"""
        entity     = request.args.get('entity')
        start_date = request.args.get('start')
        end_date   = request.args.get('end')
        limit      = int(request.args.get('limit', 100))

        q = Transaction.query.order_by(Transaction.trade_date.desc())
        broker_f      = request.args.get('broker')
        security_f    = request.args.get('security_code')
        if entity:
            q = q.filter_by(entity=entity)
        if broker_f:
            q = q.filter_by(broker=broker_f)
        if security_f:
            q = q.filter(Transaction.security_code == security_f)
        if start_date:
            q = q.filter(Transaction.trade_date >= date.fromisoformat(start_date))
        if end_date:
            q = q.filter(Transaction.trade_date <= date.fromisoformat(end_date))

        txns = q.limit(limit).all()
        return jsonify([t.to_dict() for t in txns])

    # ── Transaction history export ──────────────────────────────────────────
    @app.route('/api/transactions/export', methods=['GET'])
    def export_transactions():
        """Export transaction history as Excel"""
        entity     = request.args.get('entity')
        broker_f   = request.args.get('broker')
        limit      = int(request.args.get('limit', 9999))
        q = Transaction.query.order_by(Transaction.trade_date.desc())
        if entity:   q = q.filter_by(entity=entity)
        if broker_f: q = q.filter_by(broker=broker_f)
        txns = q.limit(limit).all()

        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        import tempfile
        wb = Workbook()
        ws = wb.active
        ws.title = '交易記錄'
        headers = ['日期','帳戶','券商','帳號','代號','名稱','股數','單價','成交金額','手續費','交易稅','淨收付']
        bold = Font(bold=True, color='FFFFFF')
        fill = PatternFill('solid', start_color='1F4E79')
        for i, h in enumerate(headers, 1):
            c = ws.cell(1, i, h)
            c.font = bold; c.fill = fill
            c.alignment = Alignment(horizontal='center')
        for r, t in enumerate(txns, 2):
            ws.cell(r,1,t.trade_date); ws.cell(r,2,t.entity); ws.cell(r,3,t.broker)
            ws.cell(r,4,t.account_no); ws.cell(r,5,t.security_code); ws.cell(r,6,t.security_name)
            ws.cell(r,7,t.shares); ws.cell(r,8,t.price); ws.cell(r,9,t.gross_amount)
            ws.cell(r,10,t.fee); ws.cell(r,11,t.tax); ws.cell(r,12,t.net_amount)
        tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        wb.save(tmp.name)
        label = f'{entity or "全部"}_{broker_f or "全券商"}'
        return send_file(tmp.name, as_attachment=True,
            download_name=f'交易記錄_{label}_{date.today().strftime("%Y%m%d")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    # ── Edit / Delete transaction ───────────────────────────────────────────
    @app.route('/api/transactions/<int:txn_id>', methods=['PUT'])
    def edit_transaction(txn_id):
        """Edit a transaction and recalculate positions"""
        txn  = Transaction.query.get_or_404(txn_id)
        data = request.json
        entity = txn.entity
        txn.trade_date    = date.fromisoformat(data['trade_date'])
        txn.security_code = data.get('security_code', txn.security_code)
        txn.security_name = data.get('security_name', txn.security_name)
        txn.shares        = float(data['shares'])
        txn.price         = float(data['price'])
        txn.gross_amount  = float(data['gross_amount'])
        txn.fee           = float(data.get('fee', txn.fee))
        txn.tax           = float(data.get('tax', txn.tax))
        txn.net_amount    = float(data['net_amount'])
        db.session.commit()
        _audit('edit', 'transactions', txn_id,
               f'編輯交易 {txn.security_code} {txn.security_name} ({txn.entity})',
               txn.to_dict())
        db.session.commit()
        recalculate_positions(entity)
        return jsonify({'status': 'ok'})

    @app.route('/api/transactions/<int:txn_id>', methods=['DELETE'])
    def delete_transaction(txn_id):
        """Delete a transaction, its linked cash entries, and recalculate."""
        txn    = Transaction.query.get_or_404(txn_id)
        entity = txn.entity
        security_code = txn.security_code

        linked_cash = CashEntry.query.filter_by(transaction_id=txn.id).all()
        affected_cash_accounts = set(ce.account_id for ce in linked_cash)

        try:
            snapshot = txn.to_dict()
            for ce in linked_cash:
                db.session.delete(ce)
            db.session.delete(txn)
            db.session.flush()

            # If security_code exists, also clean up any zero-share position directly
            if security_code:
                pos = Position.query.filter_by(entity=entity, security_code=security_code).first()
                if pos:
                    # Check if any remaining transactions exist for this entity+security
                    remaining = Transaction.query.filter_by(
                        entity=entity, security_code=security_code
                    ).filter(Transaction.shares != 0).count()
                    if remaining == 0:
                        db.session.delete(pos)

            for acct_id in affected_cash_accounts:
                _recalculate_cash_balance(acct_id)

            _audit('delete', 'transactions', txn_id,
                   f'刪除交易 {snapshot.get("security_code","?")} {snapshot.get("security_name","?")} ({snapshot.get("entity","?")}) {snapshot.get("trade_date","?")}',
                   snapshot)
            db.session.commit()
            recalculate_positions(entity)
            return jsonify({'status': 'ok', 'cash_entries_removed': len(linked_cash)})
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': str(e)}), 500

    # ── Securities lookup ────────────────────────────────────────────────────
    @app.route('/api/securities/lookup', methods=['GET'])
    def lookup_security():
        """Look up a stock code — first local DB, then TWSE"""
        code = request.args.get('code', '').strip()
        if not code:
            return jsonify({'error': 'No code provided'}), 400

        sec = Security.query.get(code)
        if sec:
            return jsonify({'code': code, 'name': sec.name, 'source': 'local'})

        name = fetch_twse_name(code)
        if name:
            return jsonify({'code': code, 'name': name, 'source': 'twse'})

        return jsonify({'code': code, 'name': None, 'source': None}), 404

    # ══════════════════════════════════════════════════════════════════════════
    # Cash routes
    # ══════════════════════════════════════════════════════════════════════════

    # ── 資金總表 ─────────────────────────────────────────────────────────────
    @app.route('/api/cash/accounts', methods=['GET'])
    def get_cash_accounts():
        """Return all cash accounts with current balances, optionally filtered."""
        entity   = request.args.get('entity')
        category = request.args.get('category')

        q = CashAccount.query.order_by(CashAccount.sort_order)
        if entity:
            q = q.filter_by(entity=entity)
        if category:
            q = q.filter_by(category=category)

        accounts = q.all()
        return jsonify([a.to_dict() for a in accounts])

    # ── 資金細項記錄 ─────────────────────────────────────────────────────────
    @app.route('/api/cash/entries', methods=['GET'])
    def get_cash_entries():
        """Return cash ledger entries, optionally filtered."""
        account_id = request.args.get('account_id')
        entity     = request.args.get('entity')
        start      = request.args.get('start')
        end        = request.args.get('end')
        limit      = int(request.args.get('limit', 200))

        q = CashEntry.query.order_by(CashEntry.entry_date.desc(), CashEntry.id.desc())

        if account_id:
            q = q.filter_by(account_id=account_id)
        elif entity:
            acct_ids = [a.id for a in CashAccount.query.filter_by(entity=entity).all()]
            q = q.filter(CashEntry.account_id.in_(acct_ids))
        if start:
            q = q.filter(CashEntry.entry_date >= date.fromisoformat(start))
        if end:
            q = q.filter(CashEntry.entry_date <= date.fromisoformat(end))

        entries = q.limit(limit).all()
        return jsonify([e.to_dict() for e in entries])


    # ── Export 資金細項 as Excel ─────────────────────────────────────────────
    @app.route('/api/cash/entries/export', methods=['GET'])
    def export_cash_entries():
        """Export 資金細項 ledger as Excel, filterable by account_id/entity/month."""
        account_id = request.args.get('account_id')
        entity     = request.args.get('entity')
        month      = request.args.get('month')  # 'YYYY-MM'

        q = CashEntry.query.order_by(CashEntry.entry_date.desc(), CashEntry.id.desc())
        if account_id:
            q = q.filter_by(account_id=account_id)
        elif entity:
            acct_ids = [a.id for a in CashAccount.query.filter_by(entity=entity).all()]
            q = q.filter(CashEntry.account_id.in_(acct_ids))
        if month:
            start = date.fromisoformat(f'{month}-01')
            if start.month == 12:
                end = date(start.year + 1, 1, 1)
            else:
                end = date(start.year, start.month + 1, 1)
            q = q.filter(CashEntry.entry_date >= start, CashEntry.entry_date < end)

        entries = q.all()
        acct_map = {a.id: a.name for a in CashAccount.query.all()}

        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        import tempfile
        wb = Workbook()
        ws = wb.active
        ws.title = '資金細項'
        headers = ['日期', '帳戶', '類型', '說明', '對方帳戶', '金額', '餘額']
        bold = Font(bold=True, color='FFFFFF')
        fill = PatternFill('solid', start_color='1F4E79')
        for i, h in enumerate(headers, 1):
            c = ws.cell(1, i, h)
            c.font = bold; c.fill = fill
            c.alignment = Alignment(horizontal='center')

        for r, e in enumerate(entries, 2):
            counterpart = '—'
            if e.linked_entry_id:
                mirror = CashEntry.query.get(e.linked_entry_id)
                if mirror:
                    counterpart = acct_map.get(mirror.account_id, mirror.account_id)
            desc = e.description or ''
            import re as _re
            desc = _re.sub(r'^\[[^\]]+→[^\]]+\]\s*', '', desc) or '—'
            ws.cell(r, 1, e.entry_date)
            ws.cell(r, 2, acct_map.get(e.account_id, e.account_id))
            ws.cell(r, 3, e.entry_type)
            ws.cell(r, 4, desc)
            ws.cell(r, 5, counterpart)
            ws.cell(r, 6, e.amount)
            ws.cell(r, 7, e.balance_after)

        tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        wb.save(tmp.name)
        label = account_id or entity or '全部'
        month_label = month or '全部月份'
        return send_file(tmp.name, as_attachment=True,
            download_name=f'資金細項_{label}_{month_label}_{date.today().strftime("%Y%m%d")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    # ── Manual cash entry ────────────────────────────────────────────────────
    @app.route('/api/cash/entry', methods=['POST'])
    def add_cash_entry():
        """Add a manual cash entry. Handles transfers, loans, and simple entries."""
        data         = request.json
        account_id   = data.get('account_id', '').strip()
        entry_date   = date.fromisoformat(data['entry_date'])
        entry_type   = data.get('entry_type', '').strip()
        description  = data.get('description', '').strip()
        amount       = float(data['amount'])   # positive=in, negative=out

        acct = CashAccount.query.get(account_id)
        if not acct:
            return jsonify({'error': f'找不到帳戶：{account_id}'}), 404
        if acct.is_static:
            return jsonify({'error': '靜態帳戶請直接更新餘額'}), 400

        # ── Validation ────────────────────────────────────────────────────
        if entry_type == '其他收入':
            if amount <= 0:
                return jsonify({'error': '其他收入請填正數金額'}), 400

        elif entry_type == '其他支出':
            if amount >= 0:
                return jsonify({'error': '其他支出請填負數金額'}), 400

        elif entry_type == '轉出賬戶':
            target_id = data.get('target_account_id', '').strip()
            if not target_id:
                return jsonify({'error': '請指定轉入帳戶'}), 400
            target = CashAccount.query.get(target_id)
            if not target:
                return jsonify({'error': f'找不到轉入帳戶：{target_id}'}), 404
            if acct.entity != target.entity:
                return jsonify({'error': '轉出賬戶只限同一實體（RC↔RC 或 華強↔華強）。跨實體請用借款'}), 400
            if account_id == target_id:
                return jsonify({'error': '轉出和轉入帳戶不能相同'}), 400
            if target.category == '基金投資':
                return jsonify({'error': '貨幣基金帳戶不接受轉帳，請轉入對應的私銀帳戶'}), 400
            if amount <= 0:
                return jsonify({'error': '轉出金額請填正數，系統會自動處理方向'}), 400

        elif entry_type in ('借款', '借款還款'):
            target_id = data.get('target_account_id', '').strip()
            if not target_id:
                return jsonify({'error': '請指定對方帳戶'}), 400
            target = CashAccount.query.get(target_id)
            if not target:
                return jsonify({'error': f'找不到對方帳戶：{target_id}'}), 404
            if acct.entity == target.entity:
                return jsonify({'error': '借款只限跨實體（RC↔華強）。同一實體請用轉出賬戶'}), 400
            if target.category == '基金投資':
                return jsonify({'error': '貨幣基金帳戶不接受借款，請選對應的私銀帳戶'}), 400
            if amount <= 0:
                return jsonify({'error': '借款/還款金額請填正數，系統會自動處理方向'}), 400

        elif entry_type in ('貸款', '貸款還款'):
            if 'private' not in acct.id:
                return jsonify({'error': '貸款功能只限私銀帳戶'}), 400

        try:
            if entry_type in ('轉出賬戶', '股東往來（借）', '股東往來（還）'):
                target  = CashAccount.query.get(data['target_account_id'])

                # 股東往來 must be cross-entity (RC↔華強 only); same-entity → use 轉出賬戶
                if entry_type in ('股東往來（借）', '股東往來（還）'):
                    if not target:
                        return jsonify({'error': '找不到目標帳戶'}), 400
                    if acct.entity == target.entity:
                        return jsonify({
                            'error': f'股東往來僅限不同 entity 之間（RC↔華強）。'
                                     f'{acct.entity} 帳戶之間的轉帳請使用「轉出賬戶」功能。'
                        }), 400

                out_amt = -abs(amount)
                in_amt  =  abs(amount)

                # For 股東往來, prefix entity names so Sophie can see who→who
                if entry_type in ('股東往來（借）', '股東往來（還）'):
                    entity_label = f'[{acct.entity}→{target.entity}]'
                    out_desc = f'{entity_label} {description}'.strip() if description else entity_label
                    in_desc  = out_desc
                else:
                    out_desc = description or f'轉入{target.name}'
                    in_desc  = description or f'來自{acct.name}'

                acct.balance += out_amt
                e_out = CashEntry(
                    account_id    = account_id,
                    entry_date    = entry_date,
                    entry_type    = entry_type,
                    description   = out_desc,
                    amount        = out_amt,
                    balance_after = acct.balance,
                )
                db.session.add(e_out)
                db.session.flush()

                target.balance += in_amt
                e_in = CashEntry(
                    account_id      = target.id,
                    entry_date      = entry_date,
                    entry_type      = entry_type,
                    description     = in_desc,
                    amount          = in_amt,
                    balance_after   = target.balance,
                    linked_entry_id = e_out.id,
                )
                db.session.add(e_in)
                db.session.flush()
                e_out.linked_entry_id = e_in.id

                db.session.commit()
                return jsonify({'status': 'ok', 'entry_id': e_out.id, 'mirror_id': e_in.id})

            else:
                acct.balance += amount
                entry = CashEntry(
                    account_id    = account_id,
                    entry_date    = entry_date,
                    entry_type    = entry_type,
                    description   = description,
                    amount        = amount,
                    balance_after = acct.balance,
                )
                db.session.add(entry)
                db.session.commit()
                return jsonify({'status': 'ok', 'entry_id': entry.id})

        except Exception as e:
            db.session.rollback()
            return jsonify({'error': str(e)}), 500

    # ── Delete cash entry ────────────────────────────────────────────────────
    @app.route('/api/cash/entry/<int:entry_id>', methods=['DELETE'])
    def delete_cash_entry(entry_id):
        """Delete a cash entry and its mirror (if linked). Recalculates balances."""
        entry = CashEntry.query.get_or_404(entry_id)

        to_delete = [entry]
        if entry.linked_entry_id:
            mirror = CashEntry.query.get(entry.linked_entry_id)
            if mirror:
                to_delete.append(mirror)

        affected = set(e.account_id for e in to_delete)

        # If this cash entry is linked to a transaction, delete the transaction too
        txn_to_delete = None
        for e in to_delete:
            if e.transaction_id:
                txn_to_delete = Transaction.query.get(e.transaction_id)
                break

        try:
            snapshots = [e.to_dict() for e in to_delete]
            # Break circular FK links first so DELETE doesn't fail
            for e in to_delete:
                e.linked_entry_id = None
            db.session.flush()
            for e in to_delete:
                db.session.delete(e)
            db.session.flush()
            # Delete linked transaction (recalculate_positions handles the rest)
            if txn_to_delete:
                txn_snap = txn_to_delete.to_dict()
                db.session.delete(txn_to_delete)
                db.session.flush()
                from .logic import recalculate_positions
                recalculate_positions(txn_snap.get('entity'))
                _audit('delete', 'transactions', txn_snap['id'],
                       f'資金細項刪除觸發：刪除交易 {txn_snap.get("security_code")} {txn_snap.get("trade_date")}',
                       txn_snap)
            for acct_id in affected:
                _recalculate_cash_balance(acct_id)
            for s in snapshots:
                _audit('delete', 'cash_entries', s['id'],
                       f'刪除資金記錄 {s.get("entry_type","?")} {s.get("entry_date","?")} {s.get("description","?")} 金額{s.get("amount","?")}',
                       s)
            db.session.commit()
            return jsonify({'status': 'ok', 'transaction_deleted': txn_to_delete is not None})
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': str(e)}), 500

    @app.route('/api/cash/entry/<int:entry_id>', methods=['PUT'])
    def edit_cash_entry(entry_id):
        """Edit a manual cash entry. If linked (transfer/loan), syncs mirror amount too."""
        entry = CashEntry.query.get_or_404(entry_id)
        if entry.is_auto:
            return jsonify({'error': '自動產生的記錄不能手動編輯'}), 400

        data = request.json
        affected = {entry.account_id}
        try:
            old_amount = entry.amount
            if 'entry_date'  in data: entry.entry_date  = date.fromisoformat(data['entry_date'])
            if 'description' in data: entry.description = data['description']
            if 'amount'      in data: entry.amount      = float(data['amount'])

            # Sync mirror entry if linked
            if 'amount' in data and entry.linked_entry_id:
                mirror = CashEntry.query.get(entry.linked_entry_id)
                if mirror:
                    mirror.amount = -float(data['amount'])  # opposite sign
                    affected.add(mirror.account_id)

            db.session.flush()
            for acct_id in affected:
                _recalculate_cash_balance(acct_id)
            db.session.commit()
            return jsonify({'status': 'ok'})
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': str(e)}), 500

    # ── Update static account balance ────────────────────────────────────────
    @app.route('/api/cash/static_balance', methods=['PUT'])
    def update_static_balance():
        """Directly update balance for static accounts (其他, 富邦建國, 元大銀行)."""
        data       = request.json
        account_id = data.get('account_id', '').strip()
        balance    = float(data['balance'])

        acct = CashAccount.query.get(account_id)
        if not acct:
            return jsonify({'error': f'找不到帳戶：{account_id}'}), 404
        if not acct.is_static:
            return jsonify({'error': '此帳戶不是靜態帳戶'}), 400

        try:
            acct.balance = balance
            db.session.commit()
            return jsonify({'status': 'ok', 'balance': balance})
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': str(e)}), 500

    # ── Fund entry (申購/贖回基金) ────────────────────────────────────────────
    @app.route('/api/cash/fund_entry', methods=['POST'])
    def add_fund_entry():
        """
        申購: Sophie fills entry_date, unit_cost, units
        贖回: Sophie fills entry_date, units (negative), redemption_amount
        Auto-creates linked CashEntry on the private account.
        """
        data        = request.json
        account_id  = data.get('account_id', '').strip()
        action      = data.get('action', '').strip()
        entry_date  = date.fromisoformat(data['entry_date'])

        fund_acct = CashAccount.query.get(account_id)
        if not fund_acct or fund_acct.category != '基金投資':
            return jsonify({'error': '無效的基金帳戶'}), 400

        linked_cash_id = FUND_CASH_LINK.get(account_id)
        cash_acct = CashAccount.query.get(linked_cash_id)
        if not cash_acct:
            return jsonify({'error': '找不到連結的現金帳戶'}), 500

        prev = FundEntry.query.filter_by(account_id=account_id)\
                              .order_by(FundEntry.id.desc()).first()
        prev_cum_amount = prev.cumulative_amount if prev else 0
        prev_cum_units  = prev.cumulative_units  if prev else 0
        prev_avg        = prev.avg_unit_cost      if prev else None

        try:
            if action == '申購':
                unit_cost  = float(data['unit_cost'])
                units      = float(data['units'])
                cost       = round(unit_cost * units) - 1
                cum_amount = prev_cum_amount + cost
                cum_units  = prev_cum_units  + units
                avg        = cum_amount / cum_units if cum_units else None

                fe = FundEntry(
                    account_id        = account_id,
                    entry_date        = entry_date,
                    cost              = cost,
                    unit_cost         = unit_cost,
                    units             = units,
                    cumulative_amount = cum_amount,
                    cumulative_units  = cum_units,
                    avg_unit_cost     = avg,
                )
                db.session.add(fe)

                cash_acct.balance -= cost
                ce = CashEntry(
                    account_id    = linked_cash_id,
                    entry_date    = entry_date,
                    entry_type    = '申購基金',
                    description   = f'申購{fund_acct.name}',
                    amount        = -cost,
                    balance_after = cash_acct.balance,
                )
                db.session.add(ce)
                db.session.flush()
                fe.linked_cash_entry_id = ce.id
                fund_acct.balance = cum_amount

                db.session.commit()
                return jsonify({'status': 'ok', 'cost': cost, 'cum_amount': cum_amount})

            elif action == '贖回':
                units             = float(data['units'])
                redemption_amount = float(data['redemption_amount'])
                unit_cost         = prev_avg
                cost              = round(unit_cost * units) if unit_cost else 0
                profit            = round(redemption_amount - abs(cost))
                cum_amount        = prev_cum_amount + cost
                cum_units         = prev_cum_units  + units
                avg               = (cum_amount / cum_units) if cum_units and cum_units > 0 else None

                fe = FundEntry(
                    account_id        = account_id,
                    entry_date        = entry_date,
                    cost              = cost,
                    unit_cost         = unit_cost,
                    units             = units,
                    cumulative_amount = cum_amount,
                    cumulative_units  = cum_units,
                    avg_unit_cost     = avg,
                    redemption_amount = redemption_amount,
                    profit            = profit,
                )
                db.session.add(fe)

                cost_return = abs(cost)
                cash_acct.balance += cost_return
                ce1 = CashEntry(
                    account_id    = linked_cash_id,
                    entry_date    = entry_date,
                    entry_type    = '贖回基金',
                    description   = f'贖回{fund_acct.name}－成本',
                    amount        = cost_return,
                    balance_after = cash_acct.balance,
                )
                db.session.add(ce1)

                cash_acct.balance += profit
                ce2 = CashEntry(
                    account_id    = linked_cash_id,
                    entry_date    = entry_date,
                    entry_type    = '贖回基金',
                    description   = f'贖回{fund_acct.name}－處分利益',
                    amount        = profit,
                    balance_after = cash_acct.balance,
                )
                db.session.add(ce2)
                db.session.flush()
                fe.linked_cash_entry_id = ce1.id
                fund_acct.balance = cum_amount

                db.session.commit()
                return jsonify({'status': 'ok', 'redemption_amount': redemption_amount,
                                'profit': profit, 'cum_amount': cum_amount})

            else:
                return jsonify({'error': '無效的 action，請用 申購 或 贖回'}), 400

        except Exception as e:
            db.session.rollback()
            return jsonify({'error': str(e)}), 500

    # ── Fund entries list ────────────────────────────────────────────────────
    @app.route('/api/cash/fund_entries', methods=['GET'])
    def get_fund_entries():
        account_id = request.args.get('account_id')
        q = FundEntry.query.order_by(FundEntry.entry_date, FundEntry.id)
        if account_id:
            q = q.filter_by(account_id=account_id)
        return jsonify([e.to_dict() for e in q.all()])

    # ── Delete fund entry ────────────────────────────────────────────────────
    @app.route('/api/cash/fund_entry/<int:entry_id>', methods=['DELETE'])
    def delete_fund_entry(entry_id):
        """Delete a fund entry and its linked cash entry. Recalculates balances."""
        fe = FundEntry.query.get_or_404(entry_id)
        affected = {fe.account_id}

        # Find linked cash entry
        linked_ce = None
        if fe.linked_cash_entry_id:
            linked_ce = CashEntry.query.get(fe.linked_cash_entry_id)
            if linked_ce:
                affected.add(linked_ce.account_id)

        try:
            if linked_ce:
                linked_ce.linked_entry_id = None
                db.session.flush()
                db.session.delete(linked_ce)
            db.session.delete(fe)
            db.session.flush()

            for acct_id in affected:
                _recalculate_cash_balance(acct_id)

            db.session.commit()
            return jsonify({'status': 'ok'})
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': str(e)}), 500

    # ── Loan summary ─────────────────────────────────────────────────────────
    @app.route('/api/cash/loan_summary', methods=['GET'])
    def get_loan_summary():
        """
        Returns 私銀貸款, 其他借款, and 私銀自有資金 for RC and 華強.
        貸款 = sum of 貸款 entries minus 貸款還款 entries on private accounts
        借款 = sum of 借款 entries minus 借款還款 entries (net position)
        私銀自有資金 = private account balance - 貸款 net
        """
        result = {}
        for entity in ['RC', '華強']:
            priv_id  = 'rc_private' if entity == 'RC' else 'hq_private'
            priv_acc = CashAccount.query.get(priv_id)

            # Net 貸款 (bank loans)
            loan_entries = CashEntry.query.filter_by(account_id=priv_id).all()
            net_loan = sum(
                e.amount for e in loan_entries
                if e.entry_type in ('貸款', '貸款還款')
            )

            # Net 借款 (inter-entity loans) — find entries for this entity
            entity_acct_ids = [a.id for a in CashAccount.query.filter_by(entity=entity).all()]
            borrow_entries = CashEntry.query.filter(
                CashEntry.account_id.in_(entity_acct_ids),
                CashEntry.entry_type.in_(('股東往來（借）', '股東往來（還）'))
            ).all()
            net_borrow = sum(e.amount for e in borrow_entries)

            # 華強 has a historical 3M loan from RC not yet in DB — hardcode the floor
            if entity == '華強' and net_borrow > -3000000:
                net_borrow = -3000000

            # 調度至私銀資金: hardcoded capital ever transferred into 私銀
            # (only changes if new funds are moved from external accounts into 私銀)
            TRANSFERRED_TO_PRIVATE = {
                'RC':   43030000,
                '華強': 13470000,
            }

            result[entity] = {
                'loan':    round(net_loan),
                'borrow':  round(net_borrow),
                'net_own': TRANSFERRED_TO_PRIVATE[entity],
            }

        return jsonify(result)

    # ── One-time seed route (5.29 data load) ─────────────────────────────────
    @app.route('/api/admin/seed_529', methods=['POST'])
    def seed_529():
        """
        One-time route to load 5.29 full dataset.
        Clears all transactions/positions and rebuilds from Excel.
        Protected by a secret key passed in the request body.
        """
        data   = request.json or {}
        secret = data.get('secret', '')
        if secret != 'vit_seed_529':
            return jsonify({'error': 'unauthorized'}), 403

        BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        MAIN_XL  = os.path.join(BASE, 'data', 'main_529.xlsx')
        EXTRA_XL = os.path.join(BASE, 'data', 'extra_529.xlsx')

        if not os.path.exists(MAIN_XL):
            return jsonify({'error': f'Excel not found: {MAIN_XL}'}), 500

        try:
            from openpyxl import load_workbook

            def safe_float(v):
                try: return float(v)
                except: return None

            def parse_rc_tuni(wb):
                ws = wb['(RC)統一進出明細TWD']
                txns = []
                for r in ws.iter_rows(values_only=True):
                    if r[0] is None or not hasattr(r[0],'year') or r[0].year < 2000: continue
                    if r[2] is None: continue
                    shares = safe_float(r[4])
                    if not shares: continue
                    txns.append({'trade_date':r[0].date(),'settle_date':r[1].date() if r[1] and hasattr(r[1],'date') else None,
                        'entity':'RC','broker':'統一','account_no':'600826',
                        'security_code':str(r[2]).strip(),'security_name':str(r[3]).strip() if r[3] else '',
                        'shares':shares,'price':safe_float(r[5]) or 0,'gross_amount':abs(safe_float(r[6]) or 0),
                        'fee':safe_float(r[7]) or 0,'tax':safe_float(r[8]) or 0,'net_amount':safe_float(r[10]) or 0})
                return txns

            def parse_rc_yuanta(wb):
                ws = wb['(RC)元大進出明細TWD']
                txns = []
                for r in ws.iter_rows(values_only=True):
                    if r[0] is None or not hasattr(r[0],'year') or r[0].year < 2000: continue
                    code = r[2]
                    if code is None or str(code).strip() in ('','匯款','資金調撥'): continue
                    shares = safe_float(r[4])
                    if not shares: continue
                    try: code_str = str(int(float(code))) if isinstance(code,(int,float)) else str(code).strip()
                    except: code_str = str(code).strip()
                    txns.append({'trade_date':r[0].date(),'settle_date':r[1].date() if r[1] and hasattr(r[1],'date') else None,
                        'entity':'RC','broker':'元大','account_no':'133376',
                        'security_code':code_str,'security_name':str(r[3]).strip() if r[3] else '',
                        'shares':shares,'price':safe_float(r[5]) or 0,'gross_amount':abs(safe_float(r[6]) or 0),
                        'fee':safe_float(r[7]) or 0,'tax':safe_float(r[8]) or 0,'net_amount':safe_float(r[10]) or 0})
                return txns

            def parse_hq_tuni(wb):
                ws = wb['(華強)統一進出明細TWD']
                txns = []
                for r in ws.iter_rows(values_only=True):
                    if r[2] is None or not hasattr(r[2],'year') or r[2].year < 2000: continue
                    code = r[3]
                    if code is None: continue
                    name = str(r[4]).strip() if r[4] else ''
                    if name == '資金調撥': continue
                    shares = safe_float(r[5])
                    if not shares: continue
                    try: code_str = str(int(float(code))) if isinstance(code,(int,float)) else str(code).strip()
                    except: code_str = str(code).strip()
                    txns.append({'trade_date':r[2].date(),'settle_date':None,
                        'entity':'華強','broker':'統一','account_no':'600885',
                        'security_code':code_str,'security_name':name,
                        'shares':shares,'price':safe_float(r[6]) or 0,'gross_amount':abs(safe_float(r[7]) or 0),
                        'fee':safe_float(r[8]) or 0,'tax':safe_float(r[9]) or 0,'net_amount':safe_float(r[12]) or 0})
                return txns

            def parse_hq_yuanta(wb):
                ws = wb['(華強)元大進出明細TWD']
                txns = []
                for r in ws.iter_rows(values_only=True):
                    if r[2] is None or not hasattr(r[2],'year') or r[2].year < 2000: continue
                    code = r[3]
                    if code is None: continue
                    name = str(r[4]).strip() if r[4] else ''
                    if name in ('資金調撥','入統一證'): continue
                    shares = safe_float(r[5])
                    if not shares: continue
                    try: code_str = str(int(float(code))) if isinstance(code,(int,float)) else str(code).strip()
                    except: code_str = str(code).strip()
                    txns.append({'trade_date':r[2].date(),'settle_date':None,
                        'entity':'華強','broker':'元大','account_no':'133311',
                        'security_code':code_str,'security_name':name,
                        'shares':shares,'price':safe_float(r[5]) or 0,'gross_amount':abs(safe_float(r[7]) or 0),
                        'fee':safe_float(r[8]) or 0,'tax':safe_float(r[9]) or 0,'net_amount':safe_float(r[13]) or 0})
                return txns

            def parse_private_rc(wb):
                """私RC — date col A (index 0), fallback to previous row's date if A is None/invalid"""
                ws = wb['(私RC)TWD']
                txns = []
                SKIP_CODES = {'買美金', '賣美金', '換匯', '匯款', '捐款'}
                last_valid_date = None
                for r in ws.iter_rows(values_only=True):
                    # Get trade date: col A if valid, else use last known date
                    if r[0] and hasattr(r[0],'year') and r[0].year > 2000:
                        last_valid_date = r[0]
                        td = r[0]
                    else:
                        td = last_valid_date  # carry forward previous date
                    if td is None: continue
                    code = r[2]
                    if code is None: continue
                    try: code_str = str(int(float(code))) if isinstance(code,(int,float)) else str(code).strip()
                    except: code_str = str(code).strip()
                    if code_str in SKIP_CODES: continue
                    shares = safe_float(r[5])
                    if not shares: continue
                    txns.append({'trade_date': td.date(),
                        'settle_date': r[1].date() if r[1] and hasattr(r[1],'date') and r[1].year > 2000 else None,
                        'entity':'私銀RC','broker':'國泰','account_no':'006439',
                        'security_code':code_str,'security_name':str(r[4]).strip() if r[4] else '',
                        'shares':shares,'price':safe_float(r[6]) or 0,'gross_amount':abs(safe_float(r[7]) or 0),
                        'fee':safe_float(r[8]) or 0,'tax':safe_float(r[9]) or 0,'net_amount':safe_float(r[12]) or 0})
                return txns

            def parse_private_hq(wb):
                """私強 — date col C (index 2) or col D (index 3), fallback to previous row's date"""
                ws = wb['(私強)進出明細TWD']
                txns = []
                SKIP_CODES = {'買美金', '賣美金', '換匯', '匯款', '貨款', '借款'}
                SKIP_NAMES = {'期初金額', '資金調撥', '動撥款', '還款', '利息', '利息支出',
                              '向RC借款', '國泰產險', '借華強', '元大交割帳戶入',
                              'RC匯入', '轉入敦南', '轉入復興(統一證交割)', '賣美金買台幣',
                              '統一交割入', '元大交割帳戶'}
                last_valid_date = None
                for r in ws.iter_rows(values_only=True):
                    if r[2] and hasattr(r[2],'year') and r[2].year > 2000:
                        last_valid_date = r[2]; td = r[2]
                    elif r[3] and hasattr(r[3],'year') and r[3].year > 2000:
                        last_valid_date = r[3]; td = r[3]
                    else:
                        td = last_valid_date
                    if td is None: continue
                    code = r[4]
                    if code is None: continue
                    try: code_str = str(int(float(code))) if isinstance(code,(int,float)) else str(code).strip()
                    except: code_str = str(code).strip()
                    if code_str in SKIP_CODES: continue
                    name = str(r[6]).strip() if r[6] else ''
                    if name in SKIP_NAMES: continue
                    shares = safe_float(r[7])
                    if not shares: continue
                    txns.append({'trade_date': td.date(),
                        'settle_date': r[3].date() if r[3] and hasattr(r[3],'date') and r[3].year > 2000 else None,
                        'entity':'私銀華強','broker':'國泰','account_no':'007065',
                        'security_code':code_str,'security_name':name,
                        'shares':shares,'price':safe_float(r[8]) or 0,'gross_amount':abs(safe_float(r[9]) or 0),
                        'fee':safe_float(r[10]) or 0,'tax':safe_float(r[11]) or 0,'net_amount':safe_float(r[14]) or 0})
                return txns

            wb_main = load_workbook(MAIN_XL, read_only=True, data_only=True)

            all_txns = []
            all_txns += parse_rc_tuni(wb_main)
            all_txns += parse_rc_yuanta(wb_main)
            all_txns += parse_hq_tuni(wb_main)
            all_txns += parse_hq_yuanta(wb_main)
            all_txns += parse_private_rc(wb_main)
            all_txns += parse_private_hq(wb_main)
            all_txns.sort(key=lambda x: x['trade_date'])

            # ── Clear old data ────────────────────────────────────────────
            CashEntry.query.filter(CashEntry.transaction_id != None).delete(synchronize_session=False)
            Transaction.query.delete(synchronize_session=False)
            Position.query.delete(synchronize_session=False)
            UploadLog.query.delete(synchronize_session=False)
            db.session.commit()

            # ── Seed securities ───────────────────────────────────────────
            for t in all_txns:
                if not Security.query.get(t['security_code']):
                    db.session.add(Security(code=t['security_code'], name=t['security_name'], region='T'))
            db.session.commit()

            # ── Insert transactions ───────────────────────────────────────
            for t in all_txns:
                db.session.add(Transaction(
                    trade_date    = t['trade_date'],
                    settle_date   = t['settle_date'],
                    entity        = t['entity'],
                    broker        = t['broker'],
                    account_no    = t['account_no'],
                    security_code = t['security_code'],
                    security_name = t['security_name'],
                    shares        = t['shares'],
                    price         = t['price'],
                    gross_amount  = t['gross_amount'],
                    fee           = t['fee'],
                    tax           = t['tax'],
                    net_amount    = t['net_amount'],
                    source_file   = '大合併_onlyTW_-20260529.xlsx',
                ))
            db.session.commit()

            # ── Recalculate positions ─────────────────────────────────────
            for entity in ['RC', '華強', '私銀RC', '私銀華強']:
                recalculate_positions(entity)

            # ── Update cash balances ──────────────────────────────────────
            BALANCES = {
                'rc_dunnan':      2718301,
                'rc_tuni':        4602588,
                'rc_yuanta':      3626,
                'rc_private':     1665087,
                'rc_fund':        26043158,
                'rc_other':       822,
                'hq_tuni':        3247670,
                'hq_yuanta':      168919,
                'hq_dunnan':      1141664,
                'hq_private':     1296511,
                'hq_fund':        0,
                'hq_huanan':      576056,
                'hq_fubon':       7061,
                'hq_yuanta_bank': 39455,
            }
            for acct_id, bal in BALANCES.items():
                acct = CashAccount.query.get(acct_id)
                if acct:
                    acct.balance         = bal
                    acct.opening_balance = bal
            db.session.commit()

            return jsonify({
                'status':       'ok',
                'transactions': Transaction.query.count(),
                'positions':    Position.query.count(),
                'securities':   Security.query.count(),
            })

        except Exception as e:
            db.session.rollback()
            import traceback
            return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500


    # ── Positions split by broker (RC/華強 → 統一 + 元大) ────────────────────
    @app.route('/api/positions/by_broker', methods=['GET'])
    def get_positions_by_broker():
        """
        Returns positions grouped by account_no for RC and 華強,
        so the frontend can show 統一 vs 元大 split.
        Structure: { 'RC統一': [...], 'RC元大': [...], '華強統一': [...], '華強元大': [...],
                     '私銀RC': [...], '私銀華強': [...] }
        """
        from .models import ACCOUNT_MAP
        # account_no → broker_key label
        BROKER_LABELS = {
            '600826': 'RC統一',
            '133376': 'RC元大',
            '600885': '華強統一',
            '133311': '華強元大',
            '006439': '私銀RC',
            '007065': '私銀華強',
        }

        # Replay transactions grouped by account_no
        result = {label: {} for label in BROKER_LABELS.values()}

        txns = (Transaction.query
                .filter(Transaction.security_code.isnot(None))
                .filter(Transaction.shares != 0)
                .filter(Transaction.account_no.in_(BROKER_LABELS.keys()))
                .order_by(Transaction.trade_date)
                .all())

        for txn in txns:
            label = BROKER_LABELS.get(txn.account_no)
            if not label: continue
            code = txn.security_code
            if code not in result[label]:
                result[label][code] = {'shares': 0.0, 'total_cost': 0.0,
                                       'security_name': txn.security_name or code}
            h = result[label][code]
            if txn.shares > 0:
                h['shares']     += txn.shares
                h['total_cost'] += txn.gross_amount + txn.fee
            else:
                if h['shares'] > 0:
                    sell_qty   = abs(txn.shares)
                    avg        = h['total_cost'] / h['shares']
                    h['shares']     -= sell_qty
                    h['total_cost'] -= avg * sell_qty
                    if h['shares'] < 0.001:
                        h['shares'] = 0
                        h['total_cost'] = 0

        # Enrich with live prices from Position table, format output
        output = {}
        for label, holdings in result.items():
            positions = []
            for code, h in holdings.items():
                if h['shares'] < 0.001: continue
                # Determine entity for Position lookup
                entity_map = {
                    'RC統一': 'RC', 'RC元大': 'RC',
                    '華強統一': '華強', '華強元大': '華強',
                    '私銀RC': '私銀RC', '私銀華強': '私銀華強',
                }
                entity = entity_map[label]
                pos = Position.query.filter_by(entity=entity, security_code=code).first()
                last_price = pos.last_price if pos else None
                price_date = pos.price_date.isoformat() if pos and pos.price_date else None
                shares     = round(h['shares'], 4)
                total_cost = round(h['total_cost'], 4)
                avg_cost   = round(total_cost / shares, 4) if shares else 0
                market_val = round(last_price * shares, 0) if last_price else None
                unreal_pnl = round(market_val - total_cost, 0) if market_val is not None else None
                positions.append({
                    'security_code': code,
                    'security_name': h['security_name'],
                    'shares':        shares,
                    'total_cost':    total_cost,
                    'avg_cost':      avg_cost,
                    'last_price':    last_price,
                    'price_date':    price_date,
                    'market_value':  market_val,
                    'unrealized_pnl': unreal_pnl,
                })
            output[label] = positions

        # Attach pending (not yet allocated) stock dividends to matching broker label
        broker_label_map = {
            ('RC', '統一'):   'RC統一',   ('RC', '元大'):   'RC元大',
            ('華強', '統一'): '華強統一', ('華強', '元大'): '華強元大',
            ('RC', '私銀國泰'):   '私銀RC',
            ('華強', '私銀國泰'): '私銀華強',
        }
        pending = StockDividend.query.filter_by(allocated=False).all()
        for sd in pending:
            label = broker_label_map.get((sd.entity, sd.broker))
            if not label or label not in output:
                continue
            for p in output[label]:
                if p['security_code'] == sd.security_code:
                    p.setdefault('pending_stock_dividends', []).append({
                        'id': sd.id,
                        'bonus_shares': sd.bonus_shares,
                        'expected_allocate_date': sd.expected_allocate_date.isoformat() if sd.expected_allocate_date else None,
                    })

        return jsonify(output)


    # ── Update security price_type (成交價 / 均價) ────────────────────────────
    @app.route('/api/securities/<code>/price_type', methods=['PUT'])
    def update_price_type(code):
        """Sophie sets whether a stock uses 成交價 or 均價 (興櫃)."""
        data  = request.json
        ptype = data.get('price_type', '成交價')
        if ptype not in ('成交價', '均價'):
            return jsonify({'error': '無效的 price_type'}), 400
        sec = Security.query.get(code)
        if not sec:
            return jsonify({'error': f'找不到股票：{code}'}), 404
        sec.price_type = ptype
        db.session.commit()
        return jsonify({'status': 'ok', 'code': code, 'price_type': ptype})

    # ── Get all securities with price_type ───────────────────────────────────
    @app.route('/api/securities', methods=['GET'])
    def get_securities():
        """Return all securities (for price_type management)."""
        secs = Security.query.order_by(Security.code).all()
        return jsonify([s.to_dict() for s in secs])

    # ── Dividend helper: resolve broker → actual Position entity ────────────────
    def _dividend_position_entity(form_entity, broker):
        """
        Sophie picks entity=RC/華強 + broker=元大/統一/私銀國泰 on the dividend form.
        For 私銀國泰, the actual Position row lives under entity='私銀RC'/'私銀華強'.
        For 元大/統一, the Position row is under the plain entity (RC/華強).
        """
        if broker == '私銀國泰':
            return '私銀RC' if form_entity == 'RC' else '私銀華強'
        return form_entity

    # ══════════════════════════════════════════════════════════════════════════
    # 現金股利 (Cash Dividends)
    # ══════════════════════════════════════════════════════════════════════════
    @app.route('/api/cash_dividends', methods=['GET'])
    def get_cash_dividends():
        entity = request.args.get('entity')
        q = CashDividend.query
        if entity:
            q = q.filter_by(entity=entity)
        rows = q.order_by(CashDividend.announce_date.desc()).all()
        return jsonify([r.to_dict() for r in rows])

    @app.route('/api/cash_dividends/shares_held', methods=['GET'])
    def get_cash_dividend_shares_held():
        """Auto-calc helper: given entity+broker+security_code, return current shares held."""
        form_entity = request.args.get('entity')
        broker      = request.args.get('broker')
        code        = request.args.get('security_code')
        if not all([form_entity, broker, code]):
            return jsonify({'error': '缺少參數'}), 400
        pos_entity = _dividend_position_entity(form_entity, broker)
        pos = Position.query.filter_by(entity=pos_entity, security_code=code).first()
        return jsonify({'shares_held': pos.shares if pos else 0})

    @app.route('/api/cash_dividends', methods=['POST'])
    def add_cash_dividend():
        data = request.json
        try:
            announce_date = date.fromisoformat(data['announce_date'])
        except Exception:
            return jsonify({'error': '日期格式錯誤'}), 400

        entity = data.get('entity')
        broker = data.get('broker')
        code   = data.get('security_code')
        if entity not in DIVIDEND_ENTITIES:
            return jsonify({'error': f'歸屬賬戶必須是 {DIVIDEND_ENTITIES}'}), 400
        if broker not in DIVIDEND_BROKERS:
            return jsonify({'error': f'存放券商必須是 {DIVIDEND_BROKERS}'}), 400

        sec = Security.query.get(code)
        if not sec:
            return jsonify({'error': f'找不到股票代號：{code}'}), 404

        shares_held = float(data.get('shares_held', 0))
        div_per_share = float(data.get('dividend_per_share', 0))
        override = data.get('total_amount_override')
        total_amount = round(float(override)) if override is not None else round(shares_held * div_per_share)

        expected_date = None
        if data.get('expected_deposit_date'):
            expected_date = date.fromisoformat(data['expected_deposit_date'])

        cd = CashDividend(
            announce_date=announce_date, entity=entity, broker=broker,
            security_code=code, shares_held=shares_held,
            dividend_per_share=div_per_share, total_amount=total_amount,
            period_note=data.get('period_note'), expected_deposit_date=expected_date,
        )
        db.session.add(cd)
        db.session.commit()
        _audit('create', 'cash_dividends', cd.id,
               f'新增現金股利 {entity} {sec.name} {total_amount}', cd.to_dict())
        return jsonify(cd.to_dict()), 201

    @app.route('/api/cash_dividends/<int:div_id>', methods=['PUT'])
    def update_cash_dividend(div_id):
        cd = CashDividend.query.get(div_id)
        if not cd:
            return jsonify({'error': '找不到記錄'}), 404
        data = request.json
        if 'announce_date' in data:
            cd.announce_date = date.fromisoformat(data['announce_date'])
        if 'entity' in data:
            cd.entity = data['entity']
        if 'broker' in data:
            cd.broker = data['broker']
        if 'security_code' in data:
            sec = Security.query.get(data['security_code'])
            if not sec:
                return jsonify({'error': f'找不到股票代號：{data["security_code"]}'}), 404
            cd.security_code = data['security_code']
        if 'shares_held' in data or 'dividend_per_share' in data:
            cd.shares_held        = float(data.get('shares_held', cd.shares_held))
            cd.dividend_per_share = float(data.get('dividend_per_share', cd.dividend_per_share))
            override = data.get('total_amount_override')
            cd.total_amount = round(float(override)) if override is not None else round(cd.shares_held * cd.dividend_per_share)
        if 'period_note' in data:
            cd.period_note = data['period_note']
        if 'expected_deposit_date' in data:
            cd.expected_deposit_date = (date.fromisoformat(data['expected_deposit_date'])
                                        if data['expected_deposit_date'] else None)
        db.session.commit()
        _audit('edit', 'cash_dividends', cd.id, '編輯現金股利', cd.to_dict())
        return jsonify(cd.to_dict())

    @app.route('/api/cash_dividends/<int:div_id>/deposit', methods=['POST'])
    def deposit_cash_dividend(div_id):
        """
        Sophie clicks '已入帳' once the cash actually arrives.
        This creates a CashEntry in the appropriate 資金細項 ledger.
        """
        cd = CashDividend.query.get(div_id)
        if not cd:
            return jsonify({'error': '找不到記錄'}), 404
        if cd.deposited:
            return jsonify({'error': '此筆已標記為入帳'}), 400

        data = request.json or {}
        account_id = data.get('account_id')
        if not account_id:
            return jsonify({'error': '請指定入帳的資金帳戶'}), 400
        acct = CashAccount.query.get(account_id)
        if not acct:
            return jsonify({'error': f'找不到帳戶：{account_id}'}), 404

        sec = Security.query.get(cd.security_code)
        entry = CashEntry(
            account_id=account_id, entry_date=date.today(), entry_type='股利收入',
            description=f'{sec.name if sec else cd.security_code} 現金股利 {cd.period_note or ""}',
            amount=cd.total_amount,
        )
        acct.balance += cd.total_amount
        entry.balance_after = acct.balance
        db.session.add(entry)
        db.session.flush()

        cd.deposited    = True
        cd.deposited_at = datetime.utcnow()
        cd.cash_entry_id = entry.id
        db.session.commit()
        _audit('edit', 'cash_dividends', cd.id, f'現金股利已入帳 金額{cd.total_amount}', cd.to_dict())
        return jsonify(cd.to_dict())

    @app.route('/api/cash_dividends/<int:div_id>', methods=['DELETE'])
    def delete_cash_dividend(div_id):
        cd = CashDividend.query.get(div_id)
        if not cd:
            return jsonify({'error': '找不到記錄'}), 404
        snap = cd.to_dict()

        # If already deposited, reverse the linked CashEntry and balance first
        if cd.deposited and cd.cash_entry_id:
            entry = CashEntry.query.get(cd.cash_entry_id)
            if entry:
                acct = CashAccount.query.get(entry.account_id)
                if acct:
                    acct.balance -= entry.amount
                    _recalculate_cash_balance(acct.id)
                db.session.delete(entry)

        db.session.delete(cd)
        db.session.commit()
        _audit('delete', 'cash_dividends', div_id, '刪除現金股利（含已入帳記錄回滾）' if snap.get('deposited') else '刪除現金股利', snap)
        return jsonify({'status': 'ok'})

    # ══════════════════════════════════════════════════════════════════════════
    # 股票股利 (Stock Dividends)
    # ══════════════════════════════════════════════════════════════════════════
    @app.route('/api/stock_dividends', methods=['GET'])
    def get_stock_dividends():
        entity = request.args.get('entity')
        q = StockDividend.query
        if entity:
            q = q.filter_by(entity=entity)
        rows = q.order_by(StockDividend.announce_date.desc()).all()
        return jsonify([r.to_dict() for r in rows])

    @app.route('/api/stock_dividends', methods=['POST'])
    def add_stock_dividend():
        data = request.json
        try:
            announce_date = date.fromisoformat(data['announce_date'])
        except Exception:
            return jsonify({'error': '日期格式錯誤'}), 400

        entity = data.get('entity')
        broker = data.get('broker')
        code   = data.get('security_code')
        if entity not in DIVIDEND_ENTITIES:
            return jsonify({'error': f'歸屬賬戶必須是 {DIVIDEND_ENTITIES}'}), 400
        if broker not in DIVIDEND_BROKERS:
            return jsonify({'error': f'存放券商必須是 {DIVIDEND_BROKERS}'}), 400

        sec = Security.query.get(code)
        if not sec:
            return jsonify({'error': f'找不到股票代號：{code}'}), 404

        shares_held = float(data.get('shares_held', 0))
        ratio       = float(data.get('dividend_ratio', 0))
        bonus_shares = round(shares_held * ratio)

        expected_date = None
        if data.get('expected_allocate_date'):
            expected_date = date.fromisoformat(data['expected_allocate_date'])

        sd = StockDividend(
            announce_date=announce_date, entity=entity, broker=broker,
            security_code=code, shares_held=shares_held,
            dividend_ratio=ratio, bonus_shares=bonus_shares,
            period_note=data.get('period_note'), expected_allocate_date=expected_date,
        )
        db.session.add(sd)
        db.session.commit()
        _audit('create', 'stock_dividends', sd.id,
               f'新增股票股利 {entity} {sec.name} {bonus_shares}股', sd.to_dict())
        return jsonify(sd.to_dict()), 201

    @app.route('/api/stock_dividends/<int:div_id>', methods=['PUT'])
    def update_stock_dividend(div_id):
        sd = StockDividend.query.get(div_id)
        if not sd:
            return jsonify({'error': '找不到記錄'}), 404
        data = request.json
        if 'announce_date' in data:
            sd.announce_date = date.fromisoformat(data['announce_date'])
        if 'entity' in data:
            sd.entity = data['entity']
        if 'broker' in data:
            sd.broker = data['broker']
        if 'security_code' in data:
            sec = Security.query.get(data['security_code'])
            if not sec:
                return jsonify({'error': f'找不到股票代號：{data["security_code"]}'}), 404
            sd.security_code = data['security_code']
        if 'shares_held' in data or 'dividend_ratio' in data:
            sd.shares_held    = float(data.get('shares_held', sd.shares_held))
            sd.dividend_ratio = float(data.get('dividend_ratio', sd.dividend_ratio))
            sd.bonus_shares   = round(sd.shares_held * sd.dividend_ratio)
        if 'period_note' in data:
            sd.period_note = data['period_note']
        if 'expected_allocate_date' in data:
            sd.expected_allocate_date = (date.fromisoformat(data['expected_allocate_date'])
                                         if data['expected_allocate_date'] else None)
        db.session.commit()
        _audit('edit', 'stock_dividends', sd.id, '編輯股票股利', sd.to_dict())
        return jsonify(sd.to_dict())

    @app.route('/api/stock_dividends/<int:div_id>/allocate', methods=['POST'])
    def allocate_stock_dividend(div_id):
        """
        Sophie clicks '已入集保' once the bonus shares are confirmed tradeable.
        Creates a zero-cost Transaction for the bonus shares, then replays
        recalculate_positions() so it survives future price updates / edits
        (Position is always derived from Transaction history, never patched directly).
        """
        sd = StockDividend.query.get(div_id)
        if not sd:
            return jsonify({'error': '找不到記錄'}), 404
        if sd.allocated:
            return jsonify({'error': '此筆已標記為入集保'}), 400

        pos_entity = _dividend_position_entity(sd.entity, sd.broker)
        sec = Security.query.get(sd.security_code)

        # account_no: reuse whatever account this entity+broker combo normally trades through
        account_no = None
        for acc_no, info in ACCOUNT_MAP.items():
            if info['entity'] == pos_entity and info['broker'] == (
                '國泰' if sd.broker == '私銀國泰' else sd.broker
            ):
                account_no = acc_no
                break
        if not account_no:
            return jsonify({'error': f'找不到對應交易帳戶：{pos_entity} {sd.broker}'}), 404

        txn = Transaction(
            trade_date=date.today(), settle_date=date.today(),
            entity=pos_entity, broker=('國泰' if sd.broker=='私銀國泰' else sd.broker),
            account_no=account_no, security_code=sd.security_code,
            security_name=sec.name if sec else sd.security_code,
            shares=sd.bonus_shares, price=0, gross_amount=0, fee=0, tax=0, net_amount=0,
            memo=f'股票股利入集保（{sd.period_note or ""}）',
        )
        db.session.add(txn)
        db.session.flush()
        recalculate_positions(pos_entity)

        sd.allocated    = True
        sd.allocated_at = datetime.utcnow()
        db.session.commit()
        _audit('edit', 'stock_dividends', sd.id,

               f'股票股利已入集保 {pos_entity} {sd.security_code} +{sd.bonus_shares}股', sd.to_dict())
        return jsonify(sd.to_dict())

    @app.route('/api/stock_dividends/<int:div_id>', methods=['DELETE'])
    def delete_stock_dividend(div_id):
        sd = StockDividend.query.get(div_id)
        if not sd:
            return jsonify({'error': '找不到記錄'}), 404
        snap = sd.to_dict()

        # If already allocated, find and remove the synthetic Transaction we created,
        # then recalculate positions so shares/cost reverse cleanly
        pos_entity_for_cleanup = None
        if sd.allocated:
            pos_entity_for_cleanup = _dividend_position_entity(sd.entity, sd.broker)
            matching_txn = (Transaction.query
                .filter_by(entity=pos_entity_for_cleanup, security_code=sd.security_code,
                          shares=sd.bonus_shares, price=0)
                .filter(Transaction.memo.like(f'股票股利入集保%{sd.period_note or ""}%'))
                .first())
            if matching_txn:
                db.session.delete(matching_txn)

        db.session.delete(sd)
        db.session.commit()
        if pos_entity_for_cleanup:
            recalculate_positions(pos_entity_for_cleanup)
            db.session.commit()
        _audit('delete', 'stock_dividends', div_id,
               '刪除股票股利（含已入集保記錄回滾）' if snap.get('allocated') else '刪除股票股利', snap)
        return jsonify({'status': 'ok'})

    # ── Audit log ────────────────────────────────────────────────────────────
    @app.route('/api/audit_log', methods=['GET'])
    def get_audit_log():
        """Return recent audit log entries (delete/edit actions)."""
        limit = int(request.args.get('limit', 100))
        logs  = AuditLog.query.order_by(AuditLog.created_at.desc()).limit(limit).all()
        return jsonify([l.to_dict() for l in logs])

    return app


# ── Helper: recalculate running balance for a cash account ─────────────────
def _recalculate_cash_balance(account_id):
    """
    Recompute balance_after for all CashEntry rows in chronological order,
    then sync CashAccount.balance. Called after delete or edit.
    """
    acct = CashAccount.query.get(account_id)
    if not acct or acct.is_static:
        return
    entries = CashEntry.query.filter_by(account_id=account_id)\
                             .order_by(CashEntry.entry_date, CashEntry.id).all()
    running = acct.opening_balance
    for e in entries:
        running += e.amount
        e.balance_after = running
    acct.balance = running



def _migrate_add_price_type():
    """Add price_type column to securities table if it doesn't exist yet."""
    try:
        db.session.execute(db.text(
            "ALTER TABLE securities ADD COLUMN IF NOT EXISTS price_type VARCHAR(10) DEFAULT '成交價'"
        ))
        db.session.commit()
    except Exception:
        db.session.rollback()


def _seed_securities():
    """Pre-populate known securities from the Excel data"""
    known = {
        '6682':  '華旭先進',
        '3006':  '晶豪科',
        '6488':  '環球晶',
        '00981A':'主動統一台股增長',
        '00910': '第一金太空衛星',
        '8828':  '厲玉暣.KY',
        '6949':  '沛爾生醫-創',
        '7717':  '萊德光電-KY',
        '4573':  '高明鐵',
        '2308':  '台達電',
        '1717':  '長興',
        '00955': '中信日本商社',
        '7769':  '鴻勁',
    }
    for code, name in known.items():
        if not Security.query.get(code):
            db.session.add(Security(code=code, name=name))
    db.session.commit()


def _audit(action, table_name, record_id, summary, snapshot_dict=None):
    """Write one row to audit_logs."""
    import json
    log = AuditLog(
        action     = action,
        table_name = table_name,
        record_id  = record_id,
        summary    = summary,
        snapshot   = json.dumps(snapshot_dict, ensure_ascii=False, default=str) if snapshot_dict else None,
    )
    db.session.add(log)


def _seed_cash_accounts():
    """Create CashAccount rows from config if they don't exist yet.
    Also patches known incorrect opening balances from earlier deploys."""
    for i, cfg in enumerate(CASH_ACCOUNTS):
        existing = CashAccount.query.get(cfg['id'])
        if not existing:
            db.session.add(CashAccount(
                id              = cfg['id'],
                entity          = cfg['entity'],
                name            = cfg['name'],
                category        = cfg['category'],
                bank            = cfg.get('bank'),
                opening_balance = cfg['opening_balance'],
                balance         = cfg['opening_balance'],
                is_static       = cfg['is_static'],
                sort_order      = i,
            ))
        else:
            # Patch rc_dunnan if it has the old incorrect opening balance
            if cfg['id'] == 'rc_dunnan' and existing.opening_balance == 1765594:
                diff = cfg['opening_balance'] - existing.opening_balance  # -380013
                existing.opening_balance = cfg['opening_balance']
                existing.balance = existing.balance + diff
    db.session.commit()


if __name__ == '__main__':
    app = create_app()
    app.run(debug=True, port=5000)
