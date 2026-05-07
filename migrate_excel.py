"""
migrate_excel.py - Fixed version
"""

import sys, os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
from datetime import datetime, date
from backend.app import create_app
from backend.models import db, Transaction, Position, Security, ACCOUNT_MAP
from backend.logic import recalculate_positions


def parse_date(val):
    if val is None: return None
    if isinstance(val, datetime): 
        d = val.date()
        # Filter out corrupt Excel dates (1900-era)
        if d.year < 2000: return None
        return d
    if isinstance(val, date):
        if val.year < 2000: return None
        return val
    return None

def safe_float(val):
    if val is None: return 0.0
    if isinstance(val, (int, float)): return float(val)
    try: return float(str(val).replace(',',''))
    except: return 0.0

def is_skip_code(val):
    skip = ['期初','合計','資金調撥','總計','損益總計','換匯','匯款','貸款','借款',
            '配息','貨幣基金','RC借','統一匯入','貨款','元大交割入','統一交割入',
            '貨幣基金收益','貨幣基金利息','國泰產險','資金','動撥款','還款',
            '利息','匯款華強','還RC','RC匯入','轉入','轉出','申購','贖回']
    s = str(val).strip()
    if not s: return True
    if str(val).startswith('='): return True
    for k in skip:
        if k in s: return True
    return False


def migrate_rc_tuni(ws):
    """RC統一: A=trade_date, C=code, E=shares, F=price, G=gross, H=fee, I=tax"""
    trades = []
    for r in range(2, ws.max_row+1):
        td = parse_date(ws.cell(r,1).value)
        code = ws.cell(r,3).value
        shares = ws.cell(r,5).value
        if not td or not code or not shares: continue
        if is_skip_code(code) or str(shares).startswith('='): continue
        try: int(str(code).replace('A','').replace('L',''))
        except: continue
        trades.append({
            'trade_date': td, 'entity': 'RC', 'broker': '統一', 'account_no': '600826',
            'security_code': str(code).strip(), 'shares': safe_float(shares),
            'price': safe_float(ws.cell(r,6).value),
            'gross_amount': abs(safe_float(ws.cell(r,7).value)),
            'fee': safe_float(ws.cell(r,8).value), 'tax': safe_float(ws.cell(r,9).value),
            'net_amount': abs(safe_float(ws.cell(r,11).value)),
        })
    return trades


def migrate_rc_yuanta(ws):
    """RC元大: A=trade_date, C=code, E=shares, J=gross, H=fee, I=tax"""
    trades = []
    for r in range(2, ws.max_row+1):
        td = parse_date(ws.cell(r,1).value)
        code = ws.cell(r,3).value
        shares = ws.cell(r,5).value
        if not td or not code or not shares: continue
        if is_skip_code(code) or str(shares).startswith('='): continue
        try: int(str(code))
        except: continue
        gross = abs(safe_float(ws.cell(r,10).value))
        shares_v = safe_float(shares)
        price_v = safe_float(ws.cell(r,6).value) if ws.cell(r,6).value else (gross/shares_v if shares_v else 0)
        trades.append({
            'trade_date': td, 'entity': 'RC', 'broker': '元大', 'account_no': '133376',
            'security_code': str(code).strip(), 'shares': shares_v, 'price': price_v,
            'gross_amount': gross, 'fee': safe_float(ws.cell(r,8).value),
            'tax': safe_float(ws.cell(r,9).value), 'net_amount': abs(safe_float(ws.cell(r,11).value)),
        })
    return trades


def migrate_hq_tuni(ws):
    """華強統一: C=trade_date, D=code, F=shares, G=price, H=gross, I=fee, J=tax"""
    trades = []
    for r in range(2, ws.max_row+1):
        td = parse_date(ws.cell(r,3).value)
        code = ws.cell(r,4).value
        shares = ws.cell(r,6).value
        if not td or not code or not shares: continue
        if is_skip_code(code) or str(shares).startswith('='): continue
        try: int(str(code).replace('A','').replace('L',''))
        except: continue
        trades.append({
            'trade_date': td, 'entity': '華強', 'broker': '統一', 'account_no': '600885',
            'security_code': str(code).strip(), 'shares': safe_float(shares),
            'price': safe_float(ws.cell(r,7).value),
            'gross_amount': abs(safe_float(ws.cell(r,8).value)),
            'fee': safe_float(ws.cell(r,9).value), 'tax': safe_float(ws.cell(r,10).value),
            'net_amount': abs(safe_float(ws.cell(r,13).value)),
        })
    return trades


def migrate_hq_yuanta(ws):
    """華強元大: C=trade_date, D=code, F=shares, K=gross, I=fee, J=tax"""
    trades = []
    for r in range(2, ws.max_row+1):
        td = parse_date(ws.cell(r,3).value)
        code = ws.cell(r,4).value
        shares = ws.cell(r,6).value
        if not td or not code or not shares: continue
        if is_skip_code(code) or str(shares).startswith('='): continue
        try: int(str(code))
        except: continue
        gross = abs(safe_float(ws.cell(r,11).value))
        shares_v = safe_float(shares)
        price_v = safe_float(ws.cell(r,7).value) if ws.cell(r,7).value else (gross/shares_v if shares_v else 0)
        trades.append({
            'trade_date': td, 'entity': '華強', 'broker': '元大', 'account_no': '133311',
            'security_code': str(code).strip(), 'shares': shares_v, 'price': price_v,
            'gross_amount': gross, 'fee': safe_float(ws.cell(r,9).value),
            'tax': safe_float(ws.cell(r,10).value), 'net_amount': abs(safe_float(ws.cell(r,12).value)),
        })
    return trades


def migrate_private_rc(ws):
    """私RC: A=trade_date (fallback to settle date B, then last known date), C=code, F=shares"""
    trades = []
    last_valid_date = None
    for r in range(2, ws.max_row+1):
        td = parse_date(ws.cell(r,1).value) or parse_date(ws.cell(r,2).value)
        if td:
            last_valid_date = td
        code = ws.cell(r,3).value
        shares = ws.cell(r,6).value
        if not code or not shares: continue
        if is_skip_code(code) or str(shares).startswith('='): continue
        try:
            code_str = str(code).strip()
            int(code_str.replace('A','').replace('L',''))
        except:
            continue
        # Use last known date if current row has no valid date
        effective_date = td or last_valid_date
        if not effective_date: continue
        trades.append({
            'trade_date': effective_date, 'entity': '私銀RC', 'broker': '國泰', 'account_no': '006439',
            'security_code': code_str, 'shares': safe_float(shares),
            'price': safe_float(ws.cell(r,7).value),
            'gross_amount': abs(safe_float(ws.cell(r,8).value)),
            'fee': safe_float(ws.cell(r,9).value), 'tax': safe_float(ws.cell(r,10).value),
            'net_amount': abs(safe_float(ws.cell(r,13).value)),
        })
    return trades


def migrate_private_hq(ws):
    """
    私強: C=trade_date (sometimes empty), D=settle_date, E=code, H=shares, I=price, J=gross, K=fee, L=tax
    FIX: use col D as fallback when col C is empty
    """
    trades = []
    for r in range(2, ws.max_row+1):
        # Use trade date (C) first, fall back to settle date (D)
        td = parse_date(ws.cell(r,3).value) or parse_date(ws.cell(r,4).value)
        code = ws.cell(r,5).value
        shares = ws.cell(r,8).value
        if not td or not code or not shares: continue
        if is_skip_code(code) or str(shares).startswith('='): continue
        # Must be a valid stock code
        code_str = str(code).strip()
        try:
            int(code_str.replace('A','').replace('L',''))
        except:
            continue
        gross = abs(safe_float(ws.cell(r,10).value))
        shares_v = safe_float(shares)
        price_v = safe_float(ws.cell(r,9).value)
        trades.append({
            'trade_date': td, 'entity': '私銀華強', 'broker': '國泰', 'account_no': '007065',
            'security_code': code_str, 'shares': shares_v, 'price': price_v,
            'gross_amount': gross, 'fee': safe_float(ws.cell(r,11).value),
            'tax': safe_float(ws.cell(r,12).value), 'net_amount': gross,
        })
    return trades


def run_migration(excel_path: str):
    app = create_app()
    with app.app_context():
        existing = Transaction.query.count()
        if existing > 0:
            confirm = input(f'DB has {existing} transactions. Clear and re-migrate? (yes/no): ')
            if confirm.lower() != 'yes':
                print('Cancelled.')
                return
            Transaction.query.delete()
            Position.query.delete()
            db.session.commit()

        wb = openpyxl.load_workbook(excel_path, data_only=True)
        sheet_handlers = [
            ('(RC)統一進出明細TWD',   migrate_rc_tuni),
            ('(RC)元大進出明細TWD',   migrate_rc_yuanta),
            ('(華強)統一進出明細TWD', migrate_hq_tuni),
            ('(華強)元大進出明細TWD', migrate_hq_yuanta),
            ('(私RC)TWD',            migrate_private_rc),
            ('(私強)進出明細TWD',     migrate_private_hq),
        ]

        total = 0
        for sheet_name, handler in sheet_handlers:
            trades = handler(wb[sheet_name])
            print(f'{sheet_name}: {len(trades)} trades')
            for t in trades:
                code = t['security_code']
                if code and not Security.query.get(code):
                    db.session.add(Security(code=code, name=code))
                db.session.add(Transaction(
                    trade_date=t['trade_date'], entity=t['entity'], broker=t['broker'],
                    account_no=t['account_no'], security_code=t['security_code'],
                    shares=t['shares'], price=t['price'], gross_amount=t['gross_amount'],
                    fee=t['fee'], tax=t['tax'], net_amount=t['net_amount'],
                    source_file=os.path.basename(excel_path),
                ))
                total += 1

        db.session.commit()
        print(f'\n✅ Migrated {total} transactions')
        print('Recalculating positions...')
        recalculate_positions()

        
        for p in Position.query.filter(Position.shares > 0).all():
            print(f'  {p.entity} | {p.security_code} | {p.shares/1000:.0f}張 | avg {p.avg_cost:.2f}')

if __name__ == '__main__':
    path = sys.argv[1] if len(sys.argv) > 1 else '/mnt/user-data/uploads/test-20260505.xlsx'
    run_migration(path)
